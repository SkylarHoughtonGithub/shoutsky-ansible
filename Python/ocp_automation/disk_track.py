#Started first draft on 1/31/2020 by Skylar Houghton (SWH)

import os, time, pyautogui, win32gui,win32con, pyscreenshot, pytesseract, cv2, np, datetime, openpyxl, shutil, fnmatch, glob
from PIL import Image
import matplotlib.pyplot as plt
from tkinter import *
from openpyxl import *

##TODO: Look into main body to declare functions in. 
##TODO: Convert when appropriate to case statement. Cleaner than elif and has default that returns error code

#--------------------------
#Change working directory to where script and images will be
os.chdir('C:\\Users\\Public\\Documents\\Scripts\\Checklist_Automation')

##Launch image
os.startfile('data.png')
##Wait for image to fully load
time.sleep(1)

#make image full screen
hwnd = win32gui.GetForegroundWindow()
win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)


##Perform screen capture on x1,y1,x2,y2 dimensions in image. Considers current monitor size
##TODO: Is there a way to variabilize the image capture? Currently has static 3 images defined. Needs to be more dynamic.
def prod_resolution():
  root = Tk()
  screen_width = root.winfo_screenwidth()
  screen_height = root.winfo_screenheight()
  if screen_width == 1536 and screen_height == 864:
    ps_t1_prod = pyscreenshot.grab(bbox=(705, 340, 790, 535 ))
    ps_t2_prod = pyscreenshot.grab(bbox=(535, 605, 640, 790))    
    ##Save snapshots to file
    ps_t1_prod.save('ps_t1_prod.jpg')
    ps_t2_prod.save('ps_t2_prod.jpg')
  elif screen_width == 1920 and screen_height == 1080:
    ps_t1_prod = pyscreenshot.grab(bbox=(620, 385, 720, 535))
    ps_t2_prod = pyscreenshot.grab(bbox=(620, 590, 710, 740))   
    ##Save snapshots to file
    ps_t1_prod.save('ps_t1_prod.jpg')
    ps_t2_prod.save('ps_t2_prod.jpg')    
  elif screen_width == 2560 and screen_height == 1440:
    ps_t1_prod = pyscreenshot.grab(bbox=(1235, 615, 1330, 650))
    ps_t2_prod = pyscreenshot.grab(bbox=(1235, 700, 1330, 735))
    ps_t3_prod = pyscreenshot.grab(bbox=(1235, 785, 1330, 815))    
    ##Save snapshots to file
    ps_t1_prod.save('ps_t1_prod.jpg')
    ps_t1_prod.save('ps_t1_prod.jpg')
    ps_t2_prod.save('ps_t2_prod.jpg')
    ps_t3_prod.save('ps_t3_prod.jpg')
  return ps_t1_prod, ps_t2_prod, ps_t3_prod

prod_resolution() 



#--------------------------
#sanitize image and convert image value(s) to string

##Initialize Image to Text program
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

#prepare list of images that need converted
cwd = os.getcwd()
compellent_prod_list = glob.glob(cwd + '\\ps*' + '*.jpg')

def conversion():
  compellent_prod_value_list = []
  for i in compellent_prod_list:
    im_value = cv2.imread(i, cv2.IMREAD_GRAYSCALE)
    im_value = pytesseract.image_to_string(im_value, config='--psm 7')
    im_value_clean = float(im_value[:-2])
  
    #delimiter to know if the value is TB or GB or MB, and convert to GB
    im_delimit = im_value[-2:]
    if im_delimit == 'MB' or im_delimit == 'mb':
      if im_value_clean <= 512:
        im_value_clean = float(0.5)
        compellent_prod_value_list.append(im_value_clean)
      elif im_value_clean >= 513:
        im_value_clean = float(1.0)
        compellent_prod_value_list.append(im_value_clean)
    if im_delimit == 'GB' or im_delimit == 'gb':
      im_value_clean = im_value_clean / 1024
      compellent_prod_value_list.append(im_value_clean)
    if im_delimit == 'TB' or im_delimit == 'tb':
      compellent_prod_value_list.append(im_value_clean)
  return compellent_prod_value_list
      
compellent_prod_value_list = conversion()

#--------------------------
#Get and format todays value

#pull todays date and sanitize data
d = datetime.datetime.today()
d = d.replace(hour=0, minute=0, second=0, microsecond=0)
#Format the current date string
d_format = d.strftime('%m%d%Y') 

#---------------------------

#--------------------------
#Begin recording data in spreadsheets

#TODO - Need to have a way to automatically create new disk tracking sheet at beginning of the month and initialize weekdays. Could also make as it goes, depending on whats better.

#Verify disk tracking sheet is the correct one for the month
d_disk_format = d.strftime('%m - %Y')
disk_track_name = glob.glob('C:/Users/Public/Documents/Scripts/Checklist_Automation/Platform*' + d_disk_format + '.xl*')
disk_track_name = disk_track_name[0]

#if len(disk_track_name) > 1:
  #TODO: Need better error checking on multiple disk tracking sheets
 # print('More than one disk tracking sheet found. Please verify the correct one and retry.')
  
#Initialize checklist file
#platform_workbook = openpyxl.load_workbook(filename=checklist_name)

#Initialize disk tracking file
disk_workbook = openpyxl.load_workbook(filename=disk_track_name)
disk_sheet = disk_workbook.active

#Error checking to make sure we are recording in the correct cell. TODO: Afer this works, need to put the OCR function in the same order as this.
prod_list = glob.glob('C:/Users/Public/Documents/Scripts/Checklist_Automation/*prod*')
  

#Get and format Excel cell date value on Disk Tracking Sheet
#TODO: Search initialized disk sheet for todays date value
#TODO: Need to consider how to handle disk tracking sheet. Should it be premade, or made as it goes? If made as goes, the previous comment is irrelevant.

#Identify row for todays tracking
for row in disk_sheet.rows:
  for cell in row:
    if cell.value == d:
      disk_date_cell = cell.coordinate
      today_disk_row = cell.row
print('Disk sheet date cell:', disk_date_cell)
print('Disk sheet date row:', today_disk_row)

#Identify columns appropriate to data type. These sections are likely to be static but for best practice sake, doing this.+
#TODO: Would be better to identify a way to grab the column by whats similar, not exactly equal to (prod t1 in, vs prod t1 ==. I was getting a none data type error doing this before.
for row in disk_sheet.rows:
  for cell in row:
    if cell.value == 'PROD T1 Used':
      xl_prod_t1 = cell.column
      xl_prod_t1_cell = cell.coordinate
print('Prod T1:' , xl_prod_t1)
print('Prod T1 Cell:' , xl_prod_t1_cell)

for row in disk_sheet.rows:
  for cell in row:
    if cell.value == 'PROD T2 Used':
      xl_prod_t2 = cell.column
      xl_prod_t2_cell = cell.coordinate
print('Prod T2:' , xl_prod_t2)
print('Prod T2 Cell:' , xl_prod_t2_cell)

for row in disk_sheet.rows:
  for cell in row:
    if cell.value == 'PROD T3 Used':
      xl_prod_t3 = cell.column
      xl_prod_t3_cell = cell.coordinate
print('Prod T3:' , xl_prod_t3)
print('Prod T3 Cell:' , xl_prod_t3_cell)

for row in disk_sheet.rows:
  for cell in row:
    if cell.value == 'DR T1 Used':
      xl_dr_t1 = cell.column
      xl_dr_t1_cell = cell.coordinate
print('DR T1:' , xl_dr_t1)
print('DR T1 Cell:' , xl_dr_t1_cell)

for row in disk_sheet.rows:
  for cell in row:
    if cell.value == 'DR T3 Used':
      xl_dr_t3 = cell.column
      xl_dr_t3_cell = cell.coordinate
print('DR T3:' , xl_dr_t3)
print('DR T3 Cell:' , xl_dr_t3_cell)

#TODO: Make ps_t1_prod variable work
disk_sheet.cell(row=today_disk_row, column=xl_prod_t1).value = compellent_prod_value_list[0]
disk_sheet.cell(row=today_disk_row, column=xl_prod_t2).value = compellent_prod_value_list[1]
disk_sheet.cell(row=today_disk_row, column=xl_prod_t3).value = compellent_prod_value_list[2]

disk_workbook.save(disk_track_name)


#--------------------------